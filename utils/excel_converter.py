import json

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import ast
from database.database import get_db


def get_excel_distribution(elective):
    with open(f'.tmp/d_{elective}.json', 'r') as f:
        data = json.load(f)
    with open(f'.tmp/s_{elective}.json', 'r') as f:
        students_data = json.load(f)
    students_priorities = {}
    for student in students_data:
        email = student['email']
        seen_priorities = set()
        priorities = {}
        for i in range(1, 6):
            course = student[f'priority_{i}']
            if course not in seen_priorities:
                priorities[course] = i
                seen_priorities.add(course)
        students_priorities[email] = priorities
    file_path = f'.tmp/distributions_{elective}.xlsx'
    wb = Workbook()
    del wb['Sheet']
    ws_stats = wb.create_sheet('Overall Statistics', 0)
    ws_stats.append(["Distribution", "Priority", "Number of Students"])
    dfs_and_costs = []
    for distribution_name, distribution_data in data.items():
        cost = round(float(distribution_name.split("Cost")[1].split(":")[1]), 2)
        df = pd.DataFrame(distribution_data)
        df['picked priority'] = [students_priorities[row['student']][row['course']] for _, row in df.iterrows()]
        dfs_and_costs.append((cost, distribution_name, df))
    dfs_and_costs.sort()
    for cost, distribution_name, df in dfs_and_costs:
        ws = wb.create_sheet(f'{distribution_name} Cost {cost}'.replace(":", "")[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        priority_counts = df['picked priority'].value_counts().sort_index()
        for priority, count in priority_counts.items():
            ws_stats.append([distribution_name.split(',')[0], priority, count])
        ws_stats.append(["", "", ""])
    wb.save(file_path)
    wb = load_workbook(file_path)
    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_path)


def get_excel_current_hum():
    db = next(get_db())
    courses = pd.read_sql_query("SELECT * FROM courses_hum", db.bind)
    students = pd.read_sql_query(
        "SELECT email, gpa, priority_1, priority_2, priority_3, priority_4, priority_5, \"group\", completed, available FROM students_hum",
        db.bind)
    constraints = pd.read_sql_query("SELECT id, course_codename, student_email FROM constraints_hum", db.bind)
    for i, row in courses.iterrows():
        row['groups'] = ';'.join(row['groups'])
        courses.at[i, 'groups'] = row['groups']
    for i, row in students.iterrows():
        for course in row['completed']:
            if course == 'nan':
                row['completed'].remove(course)
        row['completed'] = ';'.join(row['completed'])
        row['available'] = ';'.join(row['available'])
        row['group'] = ';'.join(row['group'])
        students.at[i, 'group'] = row['group']
        students.at[i, 'completed'] = row['completed']
        students.at[i, 'available'] = row['available']
    file_path = '.tmp/table_hum.xlsx'
    with pd.ExcelWriter(file_path) as writer:
        courses.to_excel(writer, sheet_name='Courses', index=False)
        students.to_excel(writer, sheet_name='Students', index=False)
        constraints.to_excel(writer, sheet_name='Constraints', index=False)
    wb = load_workbook(file_path)

    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_path)


def get_excel_current_tech():
    db = next(get_db())
    courses = pd.read_sql_query("SELECT * FROM courses_tech", db.bind)
    students = pd.read_sql_query(
        "SELECT email, gpa, priority_1, priority_2, priority_3, priority_4, priority_5, \"group\", completed, available FROM students_tech",
        db.bind)
    constraints = pd.read_sql_query("SELECT id, course_codename, student_email FROM constraints_tech", db.bind)
    for i, row in courses.iterrows():
        row['groups'] = ';'.join(row['groups'])
        courses.at[i, 'groups'] = row['groups']
    for i, row in students.iterrows():
        for course in row['completed']:
            if course == 'nan':
                row['completed'].remove(course)
        row['completed'] = ';'.join(row['completed'])
        row['available'] = ';'.join(row['available'])
        row['group'] = ';'.join(row['group'])
        students.at[i, 'group'] = row['group']
        students.at[i, 'completed'] = row['completed']
        students.at[i, 'available'] = row['available']
    file_path = '.tmp/table_tech.xlsx'
    with pd.ExcelWriter(file_path) as writer:
        courses.to_excel(writer, sheet_name='Courses', index=False)
        students.to_excel(writer, sheet_name='Students', index=False)
        constraints.to_excel(writer, sheet_name='Constraints', index=False)
    wb = load_workbook(file_path)

    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_path)


def get_excel_example_hum():
    courses = pd.DataFrame({
        'id': [1],
        'codename': ['EXAMPLE1'],
        'type': ['hum'],
        'full_name': ['EXAMPLE COURSE'],
        'short_name': ['IMPORTANT NOTE:'],
        'description': ['KEEP hum NOTATION'],
        'instructor': ['IN type COLUMN'],
        'min_overall': [0],
        'max_overall': [1],
        'low_in_group': [2],
        'high_in_group': [3],
        'max_in_group': [4],
        'groups': ['gr_TEST']
    })
    students = pd.DataFrame({
        'email': ['a.b@innopolis.university'],
        'gpa': [0.0],
        'priority_1': ['EXAMPLE1'],
        'priority_2': ['EXAMPLE2'],
        'priority_3': ['EXAMPLE3'],
        'priority_4': ['EXAMPLE4'],
        'priority_5': ['EXAMPLE5'],
        'group': ['gr_TEST'],
        'completed': [''],
        'available': [''],
    })
    constrains = pd.DataFrame({
        'course_codename': ['EXAMPLE'],
        'student_email': ['a.b@innopolis.university']
    })
    file_path = '.tmp/example_hum.xlsx'
    with pd.ExcelWriter(file_path) as writer:
        courses.to_excel(writer, sheet_name='Courses', index=False)
        students.to_excel(writer, sheet_name='Students', index=False)
        constrains.to_excel(writer, sheet_name='Constraints', index=False)
    wb = load_workbook(file_path)

    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_path)


def get_excel_example_tech():
    courses = pd.DataFrame({
        'id': [1],
        'codename': ['EXAMPLE1'],
        'type': ['tech'],
        'full_name': ['EXAMPLE COURSE'],
        'short_name': ['IMPORTANT NOTE:'],
        'description': ['KEEP tech NOTATION'],
        'instructor': ['IN type COLUMN'],
        'min_overall': [0],
        'max_overall': [1],
        'low_in_group': [2],
        'high_in_group': [3],
        'max_in_group': [4],
        'groups': ['gr_TEST']
    })
    students = pd.DataFrame({
        'email': ['a.b@innopolis.university'],
        'gpa': [0.0],
        'priority_1': ['EXAMPLE1'],
        'priority_2': ['EXAMPLE2'],
        'priority_3': ['EXAMPLE3'],
        'priority_4': ['EXAMPLE4'],
        'priority_5': ['EXAMPLE5'],
        'group': ['gr_TEST'],
        'completed': [''],
        'available': [''],
    })
    constrains = pd.DataFrame({
        'course_codename': ['EXAMPLE'],
        'student_email': ['a.b@innopolis.university']
    })
    file_path = '.tmp/example_tech.xlsx'
    with pd.ExcelWriter(file_path) as writer:
        courses.to_excel(writer, sheet_name='Courses', index=False)
        students.to_excel(writer, sheet_name='Students', index=False)
        constrains.to_excel(writer, sheet_name='Constraints', index=False)
    wb = load_workbook(file_path)

    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    wb.save(file_path)
