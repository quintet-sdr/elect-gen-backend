from openpyxl import load_workbook
import xlwt
import pandas as pd


def process_files(studentFile, courseFile):
    courseWB = load_workbook(courseFile)
    courseSheet = courseWB[courseWB.sheetnames[0]]
    coursesInfo = {}
    for i in range(2, courseSheet.max_row + 1):
        nameQuota = []
        courseName = courseSheet.cell(row=i, column=2).value
        courseQuota = courseSheet.cell(row=i, column=3).value
        coursesInfo[courseName] = courseQuota
    students_df = pd.read_excel(studentFile)
    students_df = students_df.sort_values(by='GPA', ascending=False)
    students_df.to_excel(studentFile, index=False)
    studentWB = load_workbook(studentFile)
    studentSheet = studentWB[studentWB.sheetnames[0]]
    studentsInfo = {}
    for i in range(2, studentSheet.max_row + 1):
        studentPriorities = []
        studentID = studentSheet.cell(row=i, column=1).value
        studentPriorities.append(studentSheet.cell(row=i, column=5).value)
        studentPriorities.append(studentSheet.cell(row=i, column=6).value)
        studentPriorities.append(studentSheet.cell(row=i, column=7).value)
        studentPriorities.append(studentSheet.cell(row=i, column=8).value)
        studentPriorities.append(studentSheet.cell(row=i, column=9).value)
        studentsInfo[studentID] = studentPriorities
    results = xlwt.Workbook(encoding="utf-8")
    resultsSheetResults = results.add_sheet("Results")
    resultsSheetResults.write(0, 0, "Student ID")
    resultsSheetResults.write(0, 1, "Final priority")
    resultsSheetResults.write(0, 2, "Course Result")
    resultsSheetResults.write(0, 3, "1 chosen priority")
    resultsSheetResults.write(0, 4, "2 chosen priority")
    resultsSheetResults.write(0, 5, "3 chosen priority")
    resultsSheetResults.write(0, 6, "4 chosen priority")
    resultsSheetResults.write(0, 7, "5 chosen priority")
    for key in studentsInfo:
        finalPriority = 0
        courseResult = 0
        isDistributed = False
        for value in studentsInfo[key]:
            finalPriority += 1
            if value in coursesInfo:
                if coursesInfo[value] > 0:
                    courseResult = value
                    coursesInfo[value] -= 1
                    isDistributed = True
                    break
        if not isDistributed:
            maxKey = max(coursesInfo.keys(), key=lambda k: coursesInfo[k])
            finalPriority = "#N/A"
            courseResult = maxKey
            coursesInfo[maxKey] -= 1
        resultsSheetResults.write(key, 0, key)
        resultsSheetResults.write(key, 1, finalPriority)
        resultsSheetResults.write(key, 2, courseResult)
        for i in range(3, 8):
            resultsSheetResults.write(key, i, studentsInfo[key][i - 3])
    results.save("Results.xlsx")

